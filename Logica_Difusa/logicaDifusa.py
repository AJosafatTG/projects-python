# Tapia Grande Alberto Josafat
# Mendoza Arias Diego

import openpyxl
from openpyxl.styles import Alignment

def leer_rangos_y_valores(archivo):
    """Lee los rangos y valores desde un archivo Excel."""
    wb = openpyxl.load_workbook(archivo)
    sheet = wb.active

    rangos = []
    for row in sheet.iter_rows(min_row=2, max_row=5, min_col=1, max_col=3, values_only=True):
        nombre, minimo, maximo = row
        if minimo is None or maximo is None:
            raise ValueError(f"Error: El rango {nombre} tiene un valor mínimo o máximo vacío.")
        rangos.append((nombre, minimo, maximo))

    valores = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5, values_only=True):
        if row[0] is not None:
            valores.append(row[0])

    return rangos, valores

def calcular_media(rango):
    """Calcula la media de un rango."""
    return (rango[1] + rango[2]) / 2

def funcion_membresia(valor, rango):
    """Función de membresía triangular."""
    minimo, maximo = rango[1], rango[2]
    if valor < minimo or valor > maximo:
        return 0
    elif valor == minimo or valor == maximo:
        return 1
    else:
        media = calcular_media(rango)
        if valor < media:
            return (valor - minimo) / (media - minimo)
        else:
            return (maximo - valor) / (maximo - media)

def procesar_valores(rangos, valores):
    """Procesa los valores y calcula el grado de pertenencia."""
    resultados = []
    for valor in valores:
        grados_pertenencia = []
        for rango in rangos:
            membresia = funcion_membresia(valor, rango)
            grados_pertenencia.append(membresia)
        
        # Identificar el rango con la mayor membresía
        max_grado = max(grados_pertenencia)
        indice_rango = grados_pertenencia.index(max_grado)
        nombre_rango = rangos[indice_rango][0]
        resultados.append((valor, grados_pertenencia, nombre_rango))
    return resultados

def guardar_resultados(resultados, rangos, archivo_salida):
    """Guarda los resultados en un archivo Excel con formato."""
    wb = openpyxl.Workbook()
    sheet = wb.active

    # Escribir la media (𝑥̄) en la fila 1
    sheet.cell(1, 1, "𝑥̄")
    for i, rango in enumerate(rangos, start=2):
        media = calcular_media(rango)
        sheet.cell(1, i, media)
    sheet.cell(1, len(rangos) + 2, "")

    # Escribir encabezados con los nombres de los rangos
    sheet.cell(2, 1, "T")
    for i, rango in enumerate(rangos, start=2):
        sheet.cell(2, i, rango[0])
    sheet.cell(2, len(rangos) + 2, "δ")

    # Escribir resultados
    for i, (valor, grados, nombre_rango) in enumerate(resultados, start=3):
        sheet.cell(i, 1, valor)
        for j, grado in enumerate(grados, start=2):
            sheet.cell(i, j, round(grado, 2))
        sheet.cell(i, len(grados) + 2, nombre_rango)

    # Guardar el archivo
    wb.save(archivo_salida)

def main():
    # Leer archivo de entrada
    archivo_entrada = input("Introduce el nombre del archivo de Excel (.xlsx): ")
    try:
        rangos, valores = leer_rangos_y_valores(archivo_entrada)
    except ValueError as e:
        print(e)
        return

    # Procesar los valores
    resultados = procesar_valores(rangos, valores)

    # Guardar resultados
    archivo_salida = input("Introduce el nombre del archivo de salida (ejemplo: resultados.xlsx): ")
    guardar_resultados(resultados, rangos, archivo_salida)

    print(f"El proceso ha finalizado. Los resultados han sido guardados en {archivo_salida}")

if __name__ == "__main__":
    main()